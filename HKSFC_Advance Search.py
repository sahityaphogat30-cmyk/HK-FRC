import os 
import time
import traceback
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# --- SETTINGS ---
FILE_NAME = 'Search_List.xlsx'
URL = "https://apps.sfc.hk/publicregWeb/searchByName?locale=en#"

def run_verified_script():
    script_path = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_path, FILE_NAME)

    if not os.path.exists(excel_path):
        print(f"ERROR: Cannot find {FILE_NAME} in {script_path}")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        sheet.cell(row=1, column=4).value = "HKSFC (Advanced search)"
    except PermissionError:
        print("ERROR: Close Search_List.xlsx before running the script!")
        return

    options = Options()
    options.add_experimental_option("detach", True) 

    # --- FIX 1: BYPASS POST-HISTORY DELETION POPUPS ---
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-search-engine-choice-screen")
    options.add_argument("--disable-popup-blocking")

    # --- FIX 2: CHINA-PROOF & RESOLUTION SETTINGS ---
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.page_load_strategy = 'eager' # Stops hanging on blocked trackers

    try:
        print("Launching Chrome...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    except Exception as e:
        print(f"CRITICAL ERROR: Failed to launch Chrome: {e}")
        input("Press Enter to exit...")
        return

    wait = WebDriverWait(driver, 20)

    print("\n" + "="*75)
    print(f"{'ROW':<5} | {'NAME':<25} | {'COUNT':<8} | {'EXCEL STATUS':<12}")
    print("="*75)

    try:
        for r_idx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=r_idx, column=1).value
            if not name: continue

            try:
                # Network Retry Logic for China Laptop
                retry_count = 0
                while retry_count < 3:
                    try:
                        driver.get(URL)
                        break
                    except Exception:
                        retry_count += 1
                        time.sleep(2)
                
                # Give the page an extra moment to settle after loading
                time.sleep(2)

                # 1. Click Search Icon
                search_trigger = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.btn-search")))
                driver.execute_script("arguments[0].click();", search_trigger)
                time.sleep(2)

                # 2. Click Advanced Search Accordion
                adv_accordion = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "h3.advanced-search-btn")))
                driver.execute_script("arguments[0].click();", adv_accordion)

                # 3. Enter Name into Exact Phrase Field
                # Increased to 3 seconds to let the accordion fully animate open
                time.sleep(3) 
                input_field = wait.until(EC.visibility_of_element_located((By.ID, "as_epq")))
                input_field.clear()
                input_field.send_keys(str(name))

                # 4. Click Final Search
                final_search_btn = wait.until(EC.element_to_be_clickable((By.ID, "basic-popup-search")))
                driver.execute_script("arguments[0].click();", final_search_btn)

                # 5. Result Logic
                # Increased to 5 seconds to ensure results load on slower connections
                time.sleep(5) 
                final_result = "N" 
                display_count = "0"

                count_elements = driver.find_elements(By.CLASS_NAME, "searchResultTotal")

                if count_elements:
                    raw_count = count_elements[0].text.strip()
                    display_count = "".join(filter(str.isdigit, raw_count))

                    if display_count and int(display_count) > 0:
                        final_result = "NRH"
                    else:
                        display_count = "0"

                elif "did not match any documents" in driver.page_source:
                    final_result = "N"
                    display_count = "0"

                # Update Excel (Only NRH or N)
                sheet.cell(row=r_idx, column=4).value = final_result

                # Update Script Terminal (Shows both Count and Status)
                print(f"{r_idx:<5} | {str(name)[:25]:<25} | {display_count:<8} | {final_result:<12}")

            except Exception as e:
                # Properly indented debug line to prevent crashes
                print(f"DEBUG: {e}") 
                print(f"{r_idx:<5} | {str(name)[:25]:<25} | {'ERR':<8} | {'Error':<12}")
                sheet.cell(row=r_idx, column=4).value = "Error"

            if r_idx % 3 == 0:
                try:
                    wb.save(excel_path)
                except PermissionError:
                    pass

    except BaseException as e:
        print("\n" + "!"*40)
        print("CRASH DETECTED. HERE IS THE EXACT ERROR:")
        traceback.print_exc()
        print("!"*40 + "\n")

    finally:
        try:
            wb.save(excel_path)
            print("="*75)
            print("TASK COMPLETE: Excel saved successfully.")
            print("="*75)
        except Exception as e:
            print(f"Failed to save Excel file on exit: {e}")

        try:
            driver.quit()
        except:
            pass

        input("\nPress ENTER to close this window...")

if __name__ == "__main__":
    run_verified_script()