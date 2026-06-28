import pandas as pd
import os
import time
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

class HKMAAdvancedExcelBot:
    def __init__(self):
        opts = Options()
        opts.add_experimental_option("detach", True)
        opts.add_argument("--start-maximized")
        opts.add_argument("--incognito")
        
        # Speed optimization: Disable images to load the results text faster
        prefs = {"profile.managed_default_content_settings.images": 2}
        opts.add_experimental_option("prefs", prefs)

        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), 
            options=opts
        )
        self.wait = WebDriverWait(self.driver, 20) 

    def search(self, name):
        if not name or str(name).lower() == "nan": 
            return "N"
            
        try:
            # 1. Open search page in new tab
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            self.driver.get("https://www.hkma.gov.hk/eng/other-information/advanced-search")

            # 2. Input the name
            input_box = self.wait.until(EC.element_to_be_clickable((By.ID, "exact_q")))
            input_box.clear()
            clean_name = name.strip()
            input_box.send_keys(clean_name)
            
            # 3. Click Search using ID
            search_btn = self.driver.find_element(By.ID, "form-submit")
            self.driver.execute_script("arguments[0].click();", search_btn)
            
            # 4. Wait for dynamic result count
            time.sleep(3)

            total_hits = 0
            try:
                result_el = self.wait.until(EC.visibility_of_element_located((
                    By.CSS_SELECTOR, ".result-heading .left"
                )))
                result_text = result_el.text
                match = re.search(r"of\s+(\d+)", result_text) #
                if match:
                    total_hits = int(match.group(1))
            except:
                no_result_div = self.driver.find_elements(By.ID, "no_search_result")
                if no_result_div and no_result_div[0].is_displayed():
                    total_hits = 0

            if total_hits > 0:
                status = "NRH"
                print(f"RESULT: {total_hits} hits for {clean_name} -> {status}")
            else:
                status = "N"
                print(f"RESULT: 0 hits for {clean_name} -> {status}")
                
            return status

        except Exception as e:
            print(f"Error searching {name}: {e}")
            return "ERROR"

if __name__ == "__main__":
    file_path = "Search_List.xlsx"
    target_header = "HKMA (Advanced Search )" # Exact header name
    
    if os.path.exists(file_path):
        try:
            # Load workbook using openpyxl for more control
            wb = openpyxl.load_workbook(file_path)
            sheet = wb["Sheet1"]
            
            # Find which column index has our target header
            target_col_idx = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == target_header:
                    target_col_idx = col
                    break
            
            if target_col_idx is None:
                print(f"Error: Could not find column '{target_header}' in Sheet1.")
                input("Press Enter to exit...")
                exit()

            bot = HKMAAdvancedExcelBot()
            
            # Iterate through rows starting from row 2 (skipping header)
            for row_idx in range(2, sheet.max_row + 1):
                # Read name from Column A (Column 1)
                name_val = sheet.cell(row=row_idx, column=1).value
                if name_val is None: continue
                
                res_status = bot.search(str(name_val))
                
                # Update the target column directly
                sheet.cell(row=row_idx, column=target_col_idx).value = res_status
                
                # Periodic save every 5 rows to prevent data loss
                if row_idx % 5 == 0:
                    try:
                        wb.save(file_path)
                    except PermissionError:
                        pass

            # FINAL SAVE LOGIC
            saved = False
            while not saved:
                try:
                    wb.save(file_path)
                    print("\n" + "="*45 + f"\nSUCCESS: {target_header} UPDATED\n" + "="*45)
                    saved = True
                except PermissionError:
                    print("\n⚠️  SAVE FAILED: Close 'Search_List.xlsx' immediately!")
                    input("Press ENTER once you have closed the file to retry saving...")

            print("\n[INFO] Task Complete. Windows remain open.")
            while True: time.sleep(10)
            
        except Exception as e:
            print(f"Critical Error: {e}")
            input("Press Enter to exit...")
    else:
        print(f"Error: {file_path} not found.")
print("SCRIPT_DONE")