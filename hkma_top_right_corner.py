import pandas as pd
import os
import time
import re
import sys
import urllib.parse
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

class HKMATopRightCornerBot:
    def __init__(self):
        opts = Options()
        opts.add_experimental_option("detach", True)
        opts.add_argument("--start-maximized")
        opts.add_argument("--incognito")
        
        # Speed optimization: Disable images to speed up page loads
        prefs = {"profile.managed_default_content_settings.images": 2}
        opts.add_experimental_option("prefs", prefs)

        try:
            # Initialize Chrome via WebDriver Manager
            self.driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()), 
                options=opts
            )
            # Increased wait time for dynamic content
            self.wait = WebDriverWait(self.driver, 20)
        except Exception as e:
            print(f"Error starting Chrome: {e}")
            input("Press Enter to exit...")
            sys.exit()

    def search(self, name):
        if not name or str(name).lower() == "nan": 
            return "N"
            
        try:
            # 1. Open new tab for search
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            
            # 2. Prepare the search term (No quotations)
            clean_name = name.strip()
            encoded_query = urllib.parse.quote(clean_name)
            
            # 3. Construct the direct search URL exactly as per original logic
            target_url = (
                f"https://www.hkma.gov.hk/eng/other-information/result?"
                f"query={encoded_query}&a_submit=true&ui_lang=en&r_lang=all"
                f"&ui_charset=utf-8&web=this&gp0=hkma_r3_home&gp1=hkma_r3_home"
                f"&p_size=20&last_mod=%23-1"
            )
            
            self.driver.get(target_url)
            
            # 4. Mandatory sleep to allow the dynamic 'Results' count to update 
            time.sleep(3)

            total_hits = 0
            try:
                # Targeted selector for the results heading
                result_el = self.driver.find_element(By.CSS_SELECTOR, ".result-heading .left")
                result_text = result_el.text # Expected: "Results 1 - 20 of 51."
                
                # Regex to extract the digit after 'of'
                match = re.search(r"of\s+(\d+)", result_text)
                if match:
                    total_hits = int(match.group(1))
            except:
                # Check if 'No result' is displayed
                no_res = self.driver.find_elements(By.ID, "no_search_result")
                if no_res and no_res[0].is_displayed():
                    total_hits = 0

            # 5. Logic: If results > 0 -> NRH. If results == 0 -> N.
            if total_hits > 0:
                status = "NRH"
                print(f"RESULT: Found {total_hits} hits for {clean_name} -> {status}")
            else:
                status = "N"
                print(f"RESULT: Found 0 hits for {clean_name} -> {status}")
                
            return status

        except Exception as e:
            print(f"Error searching {name}: {e}")
            return "ERROR"

if __name__ == "__main__":
    file_path = "Search_List.xlsx"
    target_header = "HKMA (Top right hand corner)" # Exact header in Column H
    
    if os.path.exists(file_path):
        try:
            # Use openpyxl to update the existing column without creating new ones
            wb = openpyxl.load_workbook(file_path)
            if "Sheet1" not in wb.sheetnames:
                print("Error: Could not find 'Sheet1' in Search_List.xlsx")
                exit()
            
            sheet = wb["Sheet1"]
            
            # Find the existing target column index
            target_col_idx = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == target_header:
                    target_col_idx = col
                    break
            
            if target_col_idx is None:
                print(f"Error: Could not find column '{target_header}' in Sheet1.")
                input("Press Enter to exit...")
                exit()

            bot = HKMATopRightCornerBot()
            
            # Iterate through names in Column A (Column 1)
            for row_idx in range(2, sheet.max_row + 1):
                name_val = sheet.cell(row=row_idx, column=1).value
                if name_val is None: continue
                
                res_status = bot.search(str(name_val))
                
                # Update the target column directly
                sheet.cell(row=row_idx, column=target_col_idx).value = res_status
                
                # Intermediate save
                if row_idx % 5 == 0:
                    try:
                        wb.save(file_path)
                    except PermissionError:
                        pass
            
            # Final save logic with permission check
            saved = False
            while not saved:
                try:
                    wb.save(file_path)
                    print("\n" + "="*45 + f"\nSUCCESS: {target_header} UPDATED\n" + "="*45)
                    saved = True
                except PermissionError:
                    print("\n⚠️  PERMISSION ERROR: Please CLOSE Search_List.xlsx!")
                    input("Press ENTER once you have closed the file to retry saving...")
            
            print("\n[INFO] Task Complete. Windows remain open.")
            while True: time.sleep(10)
            
        except Exception as e:
            print(f"CRITICAL ERROR: {e}")
            input("Press Enter to exit...")
    else:
        print(f"Error: {file_path} not found.")
        input("Press Enter to exit...")
print("SCRIPT_DONE")